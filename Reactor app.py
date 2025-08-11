
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objs as go
import plotly.express as px
import io
import math
from io import BytesIO

# Optional dependencies with graceful fallbacks
try:
    from fpdf import FPDF
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    st.warning("FPDF not available. PDF export will be disabled.")

try:
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    st.warning("OpenPyXL not available. Advanced Excel export will use basic functionality.")

# -------------------- Engineering Calculations --------------------

# Material properties lookup
MATERIAL_PROPERTIES = {
    "Carbon Steel": {"allowable_stress_MPa": 137.0, "density_kg_m3": 7850, "k_W_mK": 54.0, "E_GPa": 210},
    "SS316": {"allowable_stress_MPa": 138.0, "density_kg_m3": 8000, "k_W_mK": 16.0, "E_GPa": 200},
    "Hastelloy": {"allowable_stress_MPa": 150.0, "density_kg_m3": 8400, "k_W_mK": 11.0, "E_GPa": 200},
    "Other": {"allowable_stress_MPa": 100.0, "density_kg_m3": 7800, "k_W_mK": 20.0, "E_GPa": 200}
}

# ASME Section VIII Div 1 thin-wall formula for internal pressure
def asme_wall_thickness(P_design, D_outer, S_allow, E=1.0, CA=0.0):
    """Calculate required wall thickness per ASME Section VIII Div 1"""
    denom = (2 * S_allow * E - 1.2 * P_design)
    if denom <= 0:
        return max(0.001, CA)
    t_req = (P_design * D_outer) / denom
    return t_req + CA

# Dittus-Boelter correlation for turbulent flow
def overall_U_estimate(Re, Pr, k_fluid, D_inner):
    """Estimate overall heat transfer coefficient using Dittus-Boelter correlation"""
    if Re <= 2300:
        return 100.0  # Conservative default for laminar flow
    Nu = 0.023 * Re**0.8 * Pr**0.4
    h = Nu * k_fluid / D_inner if D_inner > 0 else 0.0
    fouling_factor = 0.0005
    wall_resistance = 0.0005
    if h <= 0:
        return 100.0
    U_overall = 1.0 / (1.0 / h + wall_resistance + fouling_factor)
    return U_overall

# Packed-bed Ergun equation pressure drop
def packed_bed_pressure_drop(L, D_particle, void_frac, rho, mu, v):
    """Calculate pressure drop through packed bed using Ergun equation"""
    if D_particle <= 0 or void_frac <= 0:
        return 0.0
    term1 = 150 * (1 - void_frac)**2 / (void_frac**3) * (mu * v / D_particle**2)
    term2 = 1.75 * (1 - void_frac) / (void_frac**3) * (rho * v**2 / D_particle)
    return (term1 + term2) * L

# Temperature-dependent allowable stress tables
ALLOWABLE_STRESS_TABLES = {
    "Carbon Steel": {20: 137.0, 100: 137.0, 200: 130.0, 300: 120.0, 400: 105.0},
    "SS316": {20: 138.0, 100: 135.0, 200: 130.0, 300: 120.0, 400: 100.0},
    "Hastelloy": {20: 150.0, 100: 148.0, 200: 145.0, 300: 140.0, 400: 130.0},
    "Other": {20: 100.0, 100: 95.0, 200: 90.0, 300: 85.0, 400: 80.0}
}

def interp_allowable_stress(material, T_design_C):
    """Interpolate allowable stress based on temperature"""
    table = ALLOWABLE_STRESS_TABLES.get(material)
    if table is None:
        return MATERIAL_PROPERTIES.get(material, {}).get('allowable_stress_MPa', 100.0)
    
    temps = sorted(table.keys())
    if T_design_C <= temps[0]:
        return table[temps[0]]
    if T_design_C >= temps[-1]:
        return table[temps[-1]]
    
    for i in range(len(temps) - 1):
        t0, t1 = temps[i], temps[i + 1]
        if t0 <= T_design_C <= t1:
            s0, s1 = table[t0], table[t1]
            return s0 + (s1 - s0) * (T_design_C - t0) / (t1 - t0)
    return table[temps[0]]

def create_excel_with_calculations(project_name, inputs_data, results_data):
    """Create Excel file with summary and calculations"""
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Summary sheet
        summary_df = pd.DataFrame(list(inputs_data.items()), columns=['Parameter', 'Value'])
        summary_df.to_excel(writer, index=False, sheet_name='Summary')
        
        # Results sheet  
        results_df = pd.DataFrame(list(results_data.items()), columns=['Parameter', 'Value'])
        results_df.to_excel(writer, index=False, sheet_name='Results')
    
    output.seek(0)
    return output

# -------------------- Streamlit UI --------------------

st.set_page_config(page_title="Reactor Check Sheet Generator", layout="wide")
st.title("⚙️ Reactor Check Sheet Generator")
st.markdown("Generate professional reactor check sheets with engineering calculations and visualizations")

# Input Section
st.header("📝 Project & Design Inputs")
col1, col2 = st.columns(2)

with col1:
    project_name = st.text_input("Project Name", value="Reactor-001")
    vendor_name = st.text_input("Vendor", value="Vendor Co")
    reactor_type = st.selectbox("Reactor Type", ["Batch", "CSTR", "PFR Tubular", "Packed-bed"])
    material = st.selectbox("Material", ["Carbon Steel", "SS316", "Hastelloy", "Other"])
    D_outer = st.number_input("Outer Diameter (m)", value=1.0, min_value=0.1, max_value=10.0)
    CA = st.number_input("Corrosion Allowance (m)", value=0.003, min_value=0.0, max_value=0.05)

with col2:
    T_design = st.number_input("Design Temperature (°C)", value=150.0, min_value=-50.0, max_value=800.0)
    P_oper = st.number_input("Operating Pressure (bar)", value=5.0, min_value=0.1, max_value=100.0)
    P_design = st.number_input("Design Pressure (bar)", value=7.0, min_value=0.1, max_value=150.0)
    flow_m3_hr = st.number_input("Flow Rate (m³/hr)", value=10.0, min_value=0.1, max_value=1000.0)
    density = st.number_input("Fluid Density (kg/m³)", value=1000.0, min_value=1.0, max_value=5000.0)
    mu = st.number_input("Viscosity (Pa·s)", value=0.001, min_value=1e-6, max_value=1.0)

# Additional parameters
st.subheader("🔧 Additional Parameters")
col3, col4 = st.columns(2)

with col3:
    Cp = st.number_input("Specific Heat Capacity (J/kg·K)", value=4000.0, min_value=500.0, max_value=10000.0)
    default_S_allow = interp_allowable_stress(material, T_design)
    st.info(f"Default allowable stress for {material} at {T_design}°C: {default_S_allow:.1f} MPa")
    S_allow = st.number_input("Allowable Stress (MPa)", value=float(default_S_allow), min_value=50.0, max_value=500.0)

with col4:
    # Performance criteria
    U_min = st.number_input("Minimum U (W/m²·K)", value=100.0, min_value=10.0, max_value=5000.0)
    DP_max = st.number_input("Maximum ΔP (Pa)", value=50000.0, min_value=1000.0, max_value=500000.0)

# -------------------- Calculations --------------------

st.header("📊 Engineering Calculations & Results")

# Basic calculations
MATERIAL_DEFAULTS = MATERIAL_PROPERTIES.get(material, {})
P_design_pa = P_design * 1e5
S_allow_pa = S_allow * 1e6

# Single point calculations
A_cross = math.pi * (D_outer**2) / 4
v = (flow_m3_hr / 3600.0) / A_cross if A_cross > 0 else 0
Re = density * v * D_outer / mu if mu > 0 else 0
Pr = Cp * mu / MATERIAL_DEFAULTS.get('k_W_mK', 0.6)
U_estimate = overall_U_estimate(Re, Pr, MATERIAL_DEFAULTS.get('k_W_mK', 0.6), D_outer)

# Pressure drop calculation
if reactor_type == "Packed-bed":
    D_particle = 0.01  # 1 cm particles
    void_frac = 0.4    # 40% void fraction
    L_bed = 3.0        # 3 m bed length
    dp_estimate = packed_bed_pressure_drop(L_bed, D_particle, void_frac, density, mu, v)
else:
    L_tube = 5.0  # 5 m equivalent length
    dp_estimate = 0.02 * (L_tube / D_outer) * 0.5 * density * v**2 if D_outer > 0 else 0

# Wall thickness
thickness = asme_wall_thickness(P_design_pa, D_outer, S_allow_pa, E=1.0, CA=CA)

# Heat transfer calculations
delta_T = 50.0  # Assumed temperature difference
Q_watts = density * (flow_m3_hr / 3600.0) * Cp * delta_T
A_ht = Q_watts / (U_estimate * delta_T) if U_estimate > 0 else 0

# Results dictionary
inputs_data = {
    "Project Name": project_name,
    "Vendor": vendor_name,
    "Reactor Type": reactor_type,
    "Material": material,
    "Design Temperature (°C)": T_design,
    "Design Pressure (bar)": P_design,
    "Flow Rate (m³/hr)": flow_m3_hr,
    "Fluid Density (kg/m³)": density,
    "Viscosity (Pa·s)": mu,
    "Allowable Stress (MPa)": S_allow
}

results_data = {
    "Wall Thickness (mm)": round(thickness * 1000, 2),
    "Cross-sectional Area (m²)": round(A_cross, 4),
    "Fluid Velocity (m/s)": round(v, 3),
    "Reynolds Number": round(Re, 0),
    "Prandtl Number": round(Pr, 3),
    "Heat Transfer Coefficient U (W/m²·K)": round(U_estimate, 1),
    "Pressure Drop (Pa)": round(dp_estimate, 1),
    "Heat Duty (kW)": round(Q_watts / 1000, 2),
    "Required Heat Transfer Area (m²)": round(A_ht, 2)
}

# Display results
col5, col6 = st.columns(2)

with col5:
    st.subheader("📋 Design Summary")
    summary_df = pd.DataFrame(list(inputs_data.items()), columns=['Parameter', 'Value'])
    st.dataframe(summary_df, use_container_width=True)

with col6:
    st.subheader("🔬 Calculated Results")
    results_df = pd.DataFrame(list(results_data.items()), columns=['Parameter', 'Value'])
    st.dataframe(results_df, use_container_width=True)

# -------------------- Visualizations --------------------

st.header("📈 Performance Analysis")

# Create flow sweep for analysis
flow_range = np.linspace(flow_m3_hr * 0.1, flow_m3_hr * 3, 50)
flow_analysis = []

for flow in flow_range:
    v_local = (flow / 3600.0) / A_cross if A_cross > 0 else 0
    Re_local = density * v_local * D_outer / mu if mu > 0 else 0
    U_local = overall_U_estimate(Re_local, Pr, MATERIAL_DEFAULTS.get('k_W_mK', 0.6), D_outer)
    
    if reactor_type == "Packed-bed":
        dp_local = packed_bed_pressure_drop(3.0, 0.01, 0.4, density, mu, v_local)
    else:
        dp_local = 0.02 * (5.0 / D_outer) * 0.5 * density * v_local**2 if D_outer > 0 else 0
    
    flow_analysis.append({
        'Flow (m³/hr)': flow,
        'U (W/m²·K)': U_local,
        'ΔP (Pa)': dp_local,
        'Velocity (m/s)': v_local
    })

analysis_df = pd.DataFrame(flow_analysis)

# Plot performance curves
col7, col8 = st.columns(2)

with col7:
    fig_u = px.line(analysis_df, x='Flow (m³/hr)', y='U (W/m²·K)', 
                    title='Heat Transfer Coefficient vs Flow Rate')
    fig_u.add_hline(y=U_min, line_dash="dash", line_color="red", 
                    annotation_text=f"Min U = {U_min}")
    fig_u.add_vline(x=flow_m3_hr, line_dash="dot", line_color="green",
                    annotation_text="Design Point")
    st.plotly_chart(fig_u, use_container_width=True)

with col8:
    fig_dp = px.line(analysis_df, x='Flow (m³/hr)', y='ΔP (Pa)',
                     title='Pressure Drop vs Flow Rate')
    fig_dp.add_hline(y=DP_max, line_dash="dash", line_color="red",
                     annotation_text=f"Max ΔP = {DP_max}")
    fig_dp.add_vline(x=flow_m3_hr, line_dash="dot", line_color="green",
                     annotation_text="Design Point")
    st.plotly_chart(fig_dp, use_container_width=True)

# Temperature-stress relationship
temp_range = np.arange(20, 401, 20)
stress_values = [interp_allowable_stress(material, t) for t in temp_range]

fig_stress = px.line(x=temp_range, y=stress_values,
                     title=f'Allowable Stress vs Temperature - {material}')
fig_stress.add_vline(x=T_design, line_dash="dot", line_color="red",
                     annotation_text=f"Design T = {T_design}°C")
fig_stress.update_xaxes(title='Temperature (°C)')
fig_stress.update_yaxes(title='Allowable Stress (MPa)')
st.plotly_chart(fig_stress, use_container_width=True)

# -------------------- Export Options --------------------

st.header("💾 Export Options")

col9, col10 = st.columns(2)

with col9:
    # Excel Export
    if st.button("📄 Generate Excel Report"):
        excel_file = create_excel_with_calculations(project_name, inputs_data, results_data)
        st.download_button(
            label="⬇️ Download Excel Report",
            data=excel_file,
            file_name=f"{project_name}_reactor_check_sheet.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

with col10:
    # PDF Export (if available)
    if PDF_AVAILABLE and st.button("📑 Generate PDF Report"):
        try:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial', 'B', 16)
            pdf.cell(0, 10, f'Reactor Check Sheet - {project_name}', 0, 1, 'C')
            pdf.ln(10)
            
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(0, 8, 'Design Inputs:', 0, 1)
            pdf.set_font('Arial', '', 10)
            for key, value in inputs_data.items():
                pdf.cell(0, 6, f'{key}: {value}', 0, 1)
            
            pdf.ln(5)
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(0, 8, 'Calculation Results:', 0, 1)
            pdf.set_font('Arial', '', 10)
            for key, value in results_data.items():
                pdf.cell(0, 6, f'{key}: {value}', 0, 1)
            
            pdf_output = pdf.output(dest='S').encode('latin-1')
            st.download_button(
                label="⬇️ Download PDF Report",
                data=pdf_output,
                file_name=f"{project_name}_reactor_check_sheet.pdf",
                mime='application/pdf'
            )
        except Exception as e:
            st.error(f"PDF generation failed: {e}")
    elif not PDF_AVAILABLE:
        st.info("PDF export requires FPDF library. Install with: pip install fpdf2")

# -------------------- Footer --------------------

st.markdown("---")
st.markdown("""
**⚠️ Important Notes:**
- All calculations are based on standard correlations and should be verified
- ASME values are representative - consult official ASME tables for final design
- This tool is for preliminary design only - full engineering review required
- Always verify material properties with supplier datasheets
""")

st.caption("Reactor Check Sheet Generator v2.0 - Professional Engineering Calculations")
